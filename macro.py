
from tkinter import *
from tkinter import ttk,filedialog,messagebox
import pickle,os,sys,csv
from pathlib import Path
import pandas as pd 
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import xlsxwriter
import os
import math
from numbers import Number
import ast

class Macro:
    SHEET_COL = ["CP/ul","Initial Volume analyzed (mL)","Final Concentrate Volume (mL)","Volume used for Extraction (ml)"
                 ,"Final RNA Extraction Volume (ul)","CP/100 ml of Sample","Detection Limit (CP/100ml)","Marker Detected?","Droplet QC Pass"]
    DDPCR_COL = ["Sample","PCRType","TargetDetected"
                 ,"DetectedNotQuantifiable","QualityControlPassed","ControlQCPass?","DropletQCPass","DetectionLowerLimit"
                 ,"N1GeneCopies","N2GeneCopies","EGeneCopies","Phi6GeneCopies"
                 ,"Comments","SampleStartTime (HHMM 24-hr)","PCRResultDate (YYMMDD)","FlowRate (in MGD)","PMMoVGeneCopies/100ml"]
    RSV_COL = ["Sample","PCRType","RSVTargetDetected","SC2TargetDetected","NVG1TargetDetected","NVG2TargetDetected"
               ,"DetectedNotQuantifiable","QualityControlPassed","ControlQCPass?","DropletQCPass","DetectionLowerLimit"
               ,"RSVGeneCopies","SC2GeneCopies","NVG1GeneCopies","NVG2GeneCopies","Phi6GeneCopies"
               ,"Comments","SampleStartTime (HHMM 24-hr)","PCRResultDate (YYMMDD)","FlowRate (in MGD)","PMMoVGeneCopies/100ml"]

    RSV_LIST = ["RSV","SC2","NVG1","NVG2","Phi6"]
    COV_LIST = ["N1","N2","Phi6"]
    wb = openpyxl.Workbook()
    red_fill = PatternFill(patternType='solid', fgColor= '00FF0000')
    green_fill = PatternFill(patternType='solid', fgColor='0039B54A')
    blue_fill = PatternFill(patternType='solid', fgColor='000000FF')

    def __init__(self, root):
        self.root = root
        self.root.title("Macro")

        self.create_widgets()

    def create_widgets(self):
        global raw_input,output_excel,entry1,entry4,cov_output,rsv_output,entry3,entry2

        raw_input = StringVar()
        output_excel = StringVar()
        # master_input = StringVar()
        cov_output = StringVar()
        rsv_output = StringVar()
        try:
            raw_input.set(pickle.load(open( "pref.dat", "rb" )))
        except:
            pass

        frm = ttk.Frame(self.root, padding=10)
        frm.pack(side=LEFT)

        # row5 = ttk.Frame(frm)
        # ttk.Button(row5, width=25, text="Load Master sheet:", command=self.inputmaster).pack(side=LEFT,padx=5)
        # entry5 = ttk.Entry(row5,width=40,textvariable=master_input)
        # entry5.config(state="readonly")
        # row5.pack(side=TOP, padx=5, pady=5)
        # entry5.pack(side=RIGHT, expand=YES, fill=X)
        # entry5.xview_moveto(1)



        row1 = ttk.Frame(frm)
        ttk.Button(row1, width=15, text="Load Raw data:", command=self.inputfile).pack(side=LEFT,padx=5)
        entry1 = ttk.Entry(row1,width=40,textvariable=raw_input)
        entry1.config(state="readonly")
        row1.pack(side=TOP, padx=5, pady=5)
        entry1.pack(side=RIGHT, expand=YES, fill=X)
        entry1.xview_moveto(1)

        # buttonrow3 = ttk.Frame(frm)
        # drop2 = OptionMenu(buttonrow3,clicked2, [])
        # drop2.pack(side=LEFT,pady=5)
        # buttonrow3.pack(side=TOP,pady=5)

        row4 = ttk.Frame(frm)
        ttk.Button(row4, width=15, text="Save Excel as:", command=self.saveresult).pack(side=LEFT,padx=5)
        entry4 = ttk.Entry(row4,width=40,textvariable=output_excel)
        entry4.config(state="readonly")
        row4.pack(side=TOP, padx=5, pady=5)
        entry4.pack(side=RIGHT, expand=YES, fill=X)

        row2 = ttk.Frame(frm)
        ttk.Button(row2, width=15, text="Save cov result as:", command=self.savecov).pack(side=LEFT,padx=5)
        entry2 = ttk.Entry(row2,width=15,textvariable=cov_output)
        entry2.config(state="readonly")
        entry2.pack(side=LEFT, expand=YES, fill=X)
        ttk.Button(row2, width=15, text="Save rsv result as:", command=self.saversv).pack(side=LEFT,padx=5)
        entry3 = ttk.Entry(row2,width=15,textvariable=rsv_output)
        entry3.config(state="readonly")
        row2.pack(side=TOP, padx=5, pady=5)

        entry3.pack(side=LEFT, expand=YES, fill=X)


        buttonrow2 = ttk.Frame(frm)
        ttk.Button(buttonrow2, text="Load", command=self.loadraw).pack(side=LEFT,padx=15)
        ttk.Button(buttonrow2, text="Run", command=self.result).pack(side=LEFT,padx=15)
        ttk.Button(buttonrow2, text="Clear", command=self.clear).pack(side=LEFT,padx=15)
        ttk.Button(buttonrow2, text="Close", command=self.close).pack(side=LEFT,padx=15)
        buttonrow2.pack(side=BOTTOM,pady=5)
        
        self.log = Text(height=20,width=50)
        self.log.pack(side=RIGHT, padx=10,pady=10)

        self.root.protocol("WM_DELETE_WINDOW", self.close)

    # def inputmaster(self):
    #     master_input.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"),("all files",
    #                                                         "*.*")]))
        
    def savecov(self):
        cov_output.set(filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("text files", "*.txt"),("all files",
                                                            "*.*")]))

        entry2.xview_moveto(1)

    def saversv(self):
        rsv_output.set(filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("text files", "*.txt"),("all files",
                                                            "*.*")]))

        entry3.xview_moveto(1)

    def inputfile(self):
        raw_input.set(filedialog.askopenfilenames(parent=self.root,filetypes=[("CSV files", "*.csv"),("all files",
                                                            "*.*")]))

        entry1.xview_moveto(1)
        
    def saveresult(self):
        output_excel.set(filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Excel files", "*.xlsx"),("all files",
                                                            "*.*")]))
        entry4.xview_moveto(1)

    def close(self):
        pickle.dump(raw_input.get(),open("pref.dat", "wb"))
        self.root.destroy()

    def logprint(self,text):
        self.log.insert(END, text + '\n')


    # def updatedropdown(self):
    #     global clicked2, drop2
    #     drop2['menu'].delete(0,'end')
    #     clicked2.set(sheet_names[0])
    #     for name in sheet_names:
    #         drop2['menu'].add_command(label=name, command=lambda name=name: clicked2.set(name))

    # def loadmaster(self):
    #     global sheet_names, wb
    #     sheet_names = []
    #     try:
    #         wb = openpyxl.load_workbook(master_input.get())
    #         sheet_names = wb.sheetnames

    #     except:
    #         self.logprint("cant open the file")
    #     if not sheet_names:
    #         self.logprint("No sheets found")
    #         return
    #     else:
    #         self.updatedropdown()
    #     self.logprint("- Loading Master sheets done !")


    def output_to_Excel(self,data,name):
        try:

            data.to_excel(name,index=False)
            
        except:
            self.logprint("- Error writing output file!")
            messagebox.showerror('Error', 'Error writing output file!') 
            return 1
        

        

    # Define a custom sort key function
    def custom_sort_key(self,sample_id):
        custom_sort_order = ["EXT", "POS", "NTC", "NEG"]
        prefix = next((p for p in custom_sort_order if sample_id.startswith(p)), None)
        if prefix is None:
            # Use a default value for sorting if prefix is not found
            return (len(custom_sort_order), sample_id)
        else:
            rest_of_string = sample_id[len(prefix):]
            return (custom_sort_order.index(prefix), rest_of_string)


    def loadraw(self):
        global paths,overall_df
        replacements = {
                ' ': '_',
                'COV': 'POS',
                'PHI': 'POS',
                'RSV': 'POS',
                'NV': 'POS',
            }

        paths= list(ast.literal_eval(raw_input.get()))
        # master_df = pd.read_excel(master_input.get(),clicked2.get())
        overall_df = pd.DataFrame()

        for file_path in paths:
            try:

                raw_df = pd.read_csv(file_path,index_col=False)
            except EXCEPTION as e :
                print(f"{e}")
        # while True:
        #     raw_df = pd.read_csv(paths,index_col=False)   
            filtered_df = raw_df[raw_df["Target"].apply(lambda x: not str(x).isnumeric())]
            if "CopiesPer20uLWell" or "AcceptedDroplets"in filtered_df.columns:
                filtered_df = filtered_df.rename(columns={"CopiesPer20uLWell": "Copies/20µLWell","AcceptedDroplets": "Accepted Droplets"})
            if "Sample description 1" in filtered_df.columns:
                filtered_df = filtered_df.rename(columns={"Sample description 1": "Sample"})
            filtered_df = filtered_df[["Sample","Target","Copies/20µLWell","Accepted Droplets","Positives"]].copy()

            for key, value in replacements.items():
                filtered_df['Sample'] = filtered_df['Sample'].str.replace(key, value)
            
            overall_df = pd.concat([overall_df,filtered_df],ignore_index=True)
        
        input_df = pd.DataFrame({"Sample": overall_df["Sample"].dropna().unique()})
        input_df["Final Concentrate Volume (mL)"] = None
        with pd.ExcelWriter(output_excel.get(), engine='openpyxl', mode='w') as writer:
            input_df.to_excel(writer,sheet_name="input",index=False)
        
        self.logprint("Done import, Go to the Excel and input Concentration.")
        try:
            os.system(f'start excel "{output_excel.get()}"')
        except Exception as e:
            print(f"An error occurred: {e}")

    def output_to_compile_sheet(self, df,df_dict):
        if not df.empty:
            df["PCRType"] = "ddPCR"
            df["DetectedNotQuantifiable"] = "No"

            df["SortKey"] = df["Sample"].apply(self.custom_sort_key)
            df = df.sort_values(by="SortKey").drop(columns="SortKey")

            for index, row in df.iterrows():
                sample_id = row["Sample"]
                droplet_mean_list = {}
                control_mean_list = {}
                for key in df_dict:
                    match_df = df_dict[key][df_dict[key]["Sample"].astype(str).str.contains(str(sample_id))]
                    if not match_df.empty:
                        droplet_mean_list[key] = match_df['Droplet QC Pass'].mean()
                        control_mean_list[key] = match_df["Marker Detected?"].mean()

                if any(value > 0.6 for value in droplet_mean_list.values()):
                    df.loc[index,"DropletQCPass"] = "Yes"
                else:
                    df.loc[index,"DropletQCPass"] = "No"

                if any(prefix in sample_id for prefix in ["EXT", "NTC", "POS", "NEG"]):

                    if "TargetDetected" in df.columns:
                        df.loc[index,"TargetDetected"] = "N/A"
                    else:
                        df.loc[index,"RSVTargetDetected"] = "N/A"
                        df.loc[index,"SC2TargetDetected"] = "N/A"
                        df.loc[index,"NVG1TargetDetected"] = "N/A"
                        df.loc[index,"NVG2TargetDetected"] = "N/A"
                    if "POS" in sample_id:
                        if any(value < 0.6 for value in control_mean_list.values()):
                            df.loc[index,"ControlQCPass?"] = "No"
                        else:
                            df.loc[index,"ControlQCPass?"] = "Yes"
                    else:
                        if any(value > 0.6 for value in control_mean_list.values()):
                            df.loc[index,"ControlQCPass?"] = "No"
                        else:
                            df.loc[index,"ControlQCPass?"] = "Yes"
                else:
                    if all((df.loc[df["Sample"].str.contains(prefix), "ControlQCPass?"] == "Yes").all() for prefix in ["EXT", "NTC", "POS", "NEG"]) and \
                        df.loc[index,["DropletQCPass"]].iloc[0] == "Yes":
                        df.loc[index,"ControlQCPass?"] = 'Yes'
                    else:
                        df.loc[index,"ControlQCPass?"] = 'No'
                    if "TargetDetected" in df.columns:
                        if any(value > 0.3 for value in control_mean_list.values()):
                            df.loc[index,"TargetDetected"] = "Yes"
                        else:
                            df.loc[index,"TargetDetected"] = "No"
                    else:
                        target_columns = ["RSV", "SC2", "NVG1", "NVG2"]

                        for column in target_columns:
                            target_detected_column = f"{column}TargetDetected"
                            
                            if column in control_mean_list and control_mean_list[column] > 0.3:
                                target_value = "Yes"
                            else:
                                target_value = "No" if column in control_mean_list else "/"
                            
                            df.loc[index, target_detected_column] = target_value
                if df.loc[index,"DropletQCPass"] == 'Yes' and df.loc[index, "ControlQCPass?"] == 'Yes':
                    df.loc[index,"QualityControlPassed"] = 'Yes'
                else:
                    df.loc[index,"QualityControlPassed"] = 'No'  
        return df

    def output_df_text(self,df):
        result_df = pd.DataFrame(columns=df.columns)
        col_index = df.columns.get_loc('Comments')
        for index, row in df.iterrows():
            sample_id = row["Sample"]
            if not any(prefix in sample_id for prefix in ["EXT", "NTC", "POS", "NEG"]):
                result_df.loc[len(result_df.index)] = row
        result_df = result_df.drop(columns=list(result_df.columns[col_index:]))
        result_df = result_df.drop(columns=["ControlQCPass?","DropletQCPass"])
        return result_df

    def check_target(self):
        try:
            wb = openpyxl.load_workbook(output_excel.get())
        except:
            self.logprint("cant open the file")

        columns_to_check = ["RSVTargetDetected","SC2TargetDetected","NVG1TargetDetected","NVG2TargetDetected", "TargetDetected"]

        for sheet_name in ["rsv", "cov"]:
            sheet = wb[sheet_name]

            column_indices = {col.internal_value: idx + 1 for idx, col in enumerate(sheet[1])}

            # Iterate through the rows and columns to check

            for col_name in columns_to_check:
                col_idx = column_indices.get(col_name)
                if col_idx is not None:
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                        cell = row[0]
                        # Check if the cell value is "Yes"
                        if cell.value == "Yes":
                            cell.fill = self.green_fill

        wb.save(output_excel.get())


    def chekc_lowerlimit(self):
        try:
            wb = openpyxl.load_workbook(output_excel.get())
        except:
            self.logprint("cant open the file")

        
        columns_to_check = ["N1GeneCopies","N2GeneCopies","RSVGeneCopies","SC2GeneCopies","NVG1GeneCopies","NVG2GeneCopies","Phi6GeneCopies"]
   
        for sheet_name in ["rsv", "cov"]:
            sheet = wb[sheet_name]

            column_indices = {col.internal_value: idx + 1 for idx, col in enumerate(sheet[1])}
            for col_name in columns_to_check:
                col_idx = column_indices.get(col_name)
                compare_idx = column_indices.get("DetectionLowerLimit")
                if col_idx is not None:
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):

                        cell = row[col_idx-1]
                        cell_compare = row[compare_idx-1]
                        if cell.value and cell_compare.value:
                            if cell.value <= cell_compare.value:
                                cell.fill = self.blue_fill

        wb.save(output_excel.get())

    def result(self):
        master_df = pd.read_excel(output_excel.get(),sheet_name="input")
        cov_result_df = pd.DataFrame(columns = self.DDPCR_COL)
        rsv_result_df = pd.DataFrame(columns = self.RSV_COL)
        df_dict_mean ={}

        target_list = overall_df["Target"].dropna().unique()
        df_dict = {target: pd.DataFrame(columns=overall_df.columns.tolist()+self.SHEET_COL) for target in target_list}
        for index, rows in overall_df.iterrows():
            if rows["Target"] in df_dict:

                # df_dict[rows["Target"]] = pd.DataFrame(rows)  
                df_dict[rows["Target"]] = pd.concat([df_dict[rows["Target"]], pd.DataFrame([rows])], ignore_index=True)
                # print(df_dict[rows["Target"]])
        # print(df_dict)
        for key in df_dict:
            for index, row in df_dict[key].iterrows():   
                master_id = row["Sample"]
                if not pd.isna(master_id):
                    matching_row_df2 = master_df[master_df['Sample'].astype(str).str.contains(str(master_id))]
                    if not matching_row_df2.empty:
                # Extract the value from the matching row in df2
                        for indexs in matching_row_df2.index:
                            concern = matching_row_df2.loc[indexs, 'Final Concentrate Volume (mL)']
                            # dilution = matching_row_df2.loc[indexs, '[Dilution factor]']
                            if not np.isnan(concern):
                                df_dict[key].loc[index, "Final Concentrate Volume (mL)"] = concern
                                # df_dict[key].loc[index,"Dilution Factor"] = dilution
                #             else:
                #                 df_dict[key].loc[index, "Final Concentrate Volume (mL)"] = 100
                # else:
                #     df_dict[key].loc[index, "Final Concentrate Volume (mL)"] = 100

            df_dict[key].sort_values(by='Sample', inplace=True)
            df_dict[key]["CP/ul"] = df_dict[key]["Copies/20µLWell"]/5
            df_dict[key]["Initial Volume analyzed (mL)"] = 100
            df_dict[key]["Volume used for Extraction (ml)"] = 0.2
            df_dict[key]["Final RNA Extraction Volume (ul)"] = 80
            df_dict[key]["Detection Limit (CP/100ml)"] = (0.6 * df_dict[key]['Final RNA Extraction Volume (ul)']) \
                * ((df_dict[key]['Final Concentrate Volume (mL)'] / df_dict[key]['Volume used for Extraction (ml)']) / df_dict[key]['Initial Volume analyzed (mL)']) * 100
            df_dict[key]["CP/100 ml of Sample"] = np.where(df_dict[key]['Positives'] >= 3, \
                (((df_dict[key]['CP/ul'] * df_dict[key]['Final RNA Extraction Volume (ul)']) * \
                (df_dict[key]['Final Concentrate Volume (mL)'] / df_dict[key]['Volume used for Extraction (ml)'])) / df_dict[key]['Initial Volume analyzed (mL)']) * 100, df_dict[key]['Detection Limit (CP/100ml)'])
            df_dict[key]["Marker Detected?"] = np.where((df_dict[key]['Accepted Droplets'] >= 8000) & (df_dict[key]['Positives'] >= 3), 1, 0)
            df_dict[key]["Droplet QC Pass"] = np.where(df_dict[key]['Accepted Droplets']>=8000,1,0)

            # print(df_dict[key]["Final Concentrate Volume (mL)"])
            result = df_dict[key].groupby('Sample')[['CP/100 ml of Sample', 'Detection Limit (CP/100ml)']].mean().reset_index()

            df_dict_mean[f'{key}_mean'] = result
            col_name = ""
            if key in self.COV_LIST:
                cov_result_df["EGeneCopies"] = 0

                if key == 'N1':
                    col_name = "N1GeneCopies"
                elif key == 'N2':
                    col_name = "N2GeneCopies"
                elif key == 'Phi6':
                    col_name = "Phi6GeneCopies"
                elif key == 'EG':
                    col_name == "EGeneCopies"
                for index, row in df_dict_mean[f'{key}_mean'].iterrows():   
                    master_id = row["Sample"]
                    if not pd.isna(master_id):
                        matching_row_df2 = cov_result_df[cov_result_df['Sample'].astype(str).str.contains(str(master_id))]
                    if not matching_row_df2.empty:
                # Extract the value from the matching row in df2
                        cov_result_df.loc[matching_row_df2.index,col_name] = row["CP/100 ml of Sample"]
                        cov_result_df.loc[matching_row_df2.index,"DetectionLowerLimit"] = row["Detection Limit (CP/100ml)"]
                    else:
                        cov_result_df.loc[len(cov_result_df.index),"Sample"] = master_id
                        cov_result_df.loc[len(cov_result_df.index)-1, col_name] = row["CP/100 ml of Sample"]
                        cov_result_df.loc[len(cov_result_df.index)-1, "DetectionLowerLimit"] = row["Detection Limit (CP/100ml)"]


            elif key in self.RSV_LIST:
                if key == 'RSV':
                    col_name = "RSVGeneCopies"
                elif key == 'SC2':
                    col_name = "SC2GeneCopies"
                elif key == 'Phi6':
                    col_name = "Phi6GeneCopies"
                elif key == 'NVG1':
                    col_name = "NVG1GeneCopies"
                elif key == 'NVG2':
                    col_name = "NVG2GeneCopies"
                for index, row in df_dict_mean[f'{key}_mean'].iterrows():   
                    master_id = row["Sample"]
                    if not pd.isna(master_id):
                        matching_row_df2 = rsv_result_df[rsv_result_df['Sample'].astype(str).str.contains(str(master_id))]
                    if not matching_row_df2.empty:
                # Extract the value from the matching row in df2
                        rsv_result_df.loc[matching_row_df2.index,col_name] = row["CP/100 ml of Sample"]
                        rsv_result_df.loc[matching_row_df2.index,"DetectionLowerLimit"] = row["Detection Limit (CP/100ml)"]
                    else:
                        rsv_result_df.loc[len(rsv_result_df.index),"Sample"] = master_id
                        rsv_result_df.loc[len(rsv_result_df.index)-1, col_name] = row["CP/100 ml of Sample"]
                        rsv_result_df.loc[len(rsv_result_df.index)-1, "DetectionLowerLimit"] = row["Detection Limit (CP/100ml)"]
                    

        # OUTPUT TO COMPILE RESULT FOR CoV
        cov_result_df = self.output_to_compile_sheet(cov_result_df,df_dict)
        rsv_result_df = self.output_to_compile_sheet(rsv_result_df,df_dict)
        # print(rsv_result_df.head)

            # Iterate through the df_dict and write each DataFrame to a new sheet
        with pd.ExcelWriter(output_excel.get(), engine='openpyxl', mode='a') as writer:
            cov_result_df.to_excel(writer,sheet_name="cov",index=False)
            rsv_result_df.to_excel(writer,sheet_name="rsv",index=False)
            for key, df in df_dict.items():
                df.to_excel(writer, sheet_name=key, index=False)
        
        df = self.output_df_text(cov_result_df)
        if cov_output.get():
            df.to_csv(cov_output.get(),sep='\t', index=False)
        df = self.output_df_text(rsv_result_df)
        if rsv_output.get():
            df.to_csv(rsv_output.get(),sep='\t', index=False)


        self.check_target()
        self.chekc_lowerlimit()

        self.logprint("Done output")

    def clear(self):
        global sheet_names
        raw_input.set("")
        output_excel.set("")
        # master_input.set("")
        cov_output.set("")
        rsv_output.set("")
        sheet_names = []
        self.log.delete(1.0, END) 
        
if __name__ == '__main__':
    root = Tk()
    app = Macro(root)
    root.mainloop()
