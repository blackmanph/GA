
from tkinter import *
from tkinter import ttk,filedialog,messagebox
import pickle,os
from pathlib import Path
import pandas as pd 
import openpyxl
import requests
import re
from io import StringIO 
from dateutil.relativedelta import relativedelta
from tkcalendar import DateEntry
from datetime import date

class VariantSpotterApp:
    end_date = date.today()
    start_date = end_date - relativedelta(months=6)
    base_url = 'https://lapis.cov-spectrum.org/open/v1/sample/' 
    mutation_url = None
    lineage_url = 'aa-mutations?pangoLineage='
    format_fild = '&dataFormat=csv'
    search_words = ['Sequences','Unique Sequence']
    workbook = None
    sheet = None
    counter =0
    us_states = [
        "Alabama", "Alaska", "Arizona", "Arkansas", "California",
        "Colorado", "Connecticut", "Delaware", "Florida", "Georgia",
        "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa",
        "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland",
        "Massachusetts", "Michigan", "Minnesota", "Mississippi", "Missouri",
        "Montana", "Nebraska", "Nevada", "New Hampshire", "New Jersey",
        "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio",
        "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
        "Tennessee", "Texas", "Utah", "Vermont", "Virginia",
        "Washington", "West Virginia", "Wisconsin", "Wyoming"
    ]

    world_region = [
        "Africa",
        "Asia",
        "Europe",
        "North America",
        "South America",
        "Antarctica",
        "Oceania"
    ]

    def __init__(self, root):
        self.root = root
        self.root.title("Variant spotter")

        self.create_widgets()

    def create_widgets(self):
        global raw_input,result_file,region,threshold,country,division,entry1,entry4
        raw_input = StringVar()
        result_file = StringVar()
        region = StringVar()
        threshold = StringVar()
        threshold.set(100)
        country = StringVar()
        division = StringVar()

        frm = ttk.Frame(self.root, padding=10)
        frm.pack(side=LEFT)

        row1 = ttk.Frame(frm)
        ttk.Button(row1, width=25, text="Setting", command=self.setting).pack(side=LEFT, padx=5)
        row1.pack(side=TOP, padx=5, pady=5)

        row2 = ttk.Frame(frm)
        ttk.Button(row2, width=25, text="Load Raw data:", command=self.inputfile).pack(side=LEFT, padx=5)
        entry1 = ttk.Entry(row2, width=40, textvariable=raw_input)
        entry1.config(state="readonly")
        row2.pack(side=TOP, padx=5, pady=5)
        entry1.pack(side=RIGHT, expand=YES, fill=X)
        entry1.xview_moveto(1)

        row4 = ttk.Frame(frm)
        ttk.Button(row4, width=25, text="Save results as:", command=self.saveresult).pack(side=LEFT, padx=5)
        entry4 = ttk.Entry(row4, width=40, textvariable=result_file)
        entry4.config(state="readonly")
        row4.pack(side=TOP, padx=5, pady=5)
        entry4.pack(side=RIGHT, expand=YES, fill=X)

        buttonrow2 = ttk.Frame(frm)
        ttk.Button(buttonrow2, text="Run", command=self.result).pack(side=LEFT, padx=15)
        ttk.Button(buttonrow2, text="Clear", command=self.clear).pack(side=LEFT, padx=15)
        ttk.Button(buttonrow2, text="Close", command=self.close).pack(side=LEFT, padx=15)
        buttonrow2.pack(side=BOTTOM, pady=5)

        self.log = Text(height=20, width=50)
        self.log.pack(side=RIGHT, padx=10, pady=10)

        self.root.protocol("WM_DELETE_WINDOW", self.close)


    def inputfile(self):
        raw_input.set(filedialog.askdirectory())
        entry1.xview_moveto(1)

    def saveresult(self):
        result_file.set(filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("Excel files", "*.xlsx"),("all files",
                                                            "*.*")]))
        entry4.xview_moveto(1)

    def clear(self):
        raw_input.set("")
        result_file.set("")
        counter=0
        self.log.delete(1.0, END) 

    def close(self):
        pickle.dump(raw_input.get(),open("pref.dat", "wb"))
        root.destroy()

    def logprint(self,text):
        self.log.insert(END, text + '\n')

    def finddiff(self,org,after):
        extra_characters = []
        for char in org:
            if char not in after:
                extra_characters.append(char)
        
        return extra_characters


    def check_del(self,alist):
        pattern = r'([A-Za-z]+)(\d+-?\d*)([A-Za-z]+)del$'
        for element in alist.copy():
            # print(element)
            if 'del' in element:
                match = re.search(pattern,element)
                if match:
                    orginal = match.group(1)
                    numbers = match.group(2)
                    after = match.group(3)
                    mismatch = self.finddiff(orginal,after)
                    if mismatch:
                        if '-' in numbers:
                            number_seg = numbers.split('-')
                            end_num = int(number_seg[1])
                            start_num = int(number_seg[0])
                            mislist = list(mismatch)
                            for i in range(1, end_num - start_num + 1):
                                new_element = f'S:{mislist[i-1]}{start_num+ i}-'
                                alist.append(new_element)
                    else:
                        new_element = f'S:{orginal}{numbers}-'
                        alist.append(new_element)    
                alist.remove(element)
        return alist
    
    def handle_df(self,df):
        global counter,threshold
        total_count = df['Count'].sum()
        # iterate through sequences
        for row in df.iterrows():
            sequence = row[1]['Sequences']
            cell = sheet[f'F{counter}']
            cell.value = sequence
            count = row[1]['Count']
            aamutation_list = self.parse(sequence)
            aamutation_list = self.check_del(aamutation_list)
            # print(aamutation_list)
            api = self.generateURL(aamutation_list,mutation_url,self.format_fild)
            # print(api)
            response = self.API_call(api)
            if response.status_code == 200:
            # Parse the CSV data from the response content
                csv_data = response.text
                
                # Use StringIO to convert the CSV data into a file-like object
                csv_file = StringIO(csv_data)
                
                # Create a CSV reader and iterate through the rows
                lineage_df = pd.read_csv(csv_file)
                if(region.get()):
                    selected_region = region.get()
                    lineage_df = lineage_df[lineage_df['region'] == selected_region]
                if not lineage_df.empty:
                    lineage_sorted_df = lineage_df.sort_values(by='count', ascending=False)
                    total_count2 = lineage_df['count'].sum()
                    lineage_sorted_df['Abundance'] = lineage_sorted_df['count']/total_count2
                    matchstr = ""
                    mismatch_dic = {}
                    for index, rows in lineage_sorted_df.iterrows():
                        # only checking the abundance of lineage that is greater than 5%
                        if str(rows['pangoLineage']).strip().lower() != 'nan':
                            if rows['count'] > int(threshold.get()):

                                api = self.generateURL(rows['pangoLineage'],self.lineage_url,self.format_fild)
                                # print(rows['pangoLineage'])
                                response = self.API_call(api)
                                if response.status_code == 200:
                                    # Parse the CSV data from the response content
                                    csv_data = response.text
                                    
                                    # Use StringIO to convert the CSV data into a file-like object
                                    csv_file = StringIO(csv_data)
                                    
                                    # Create a CSV reader and iterate through the rows
                                    mutation_df = pd.read_csv(csv_file)
                                    result_list = self.check_mutation(aamutation_list,mutation_df)
                                if(not result_list):
                                    # print(f'Sequence:{indexs+1} {rows["pangoLineage"]} Percentage:{count/total_count*100:.1f}')
                                    matchstr = f'{rows["pangoLineage"]}/{matchstr}'
                                else:
                                    # print(f'Sequence:{indexs+1} Possible lineage: {rows["pangoLineage"]} Percentatge:{count/total_count*100:.1f}')
                                    mismatch_dic[rows["pangoLineage"]] = result_list
                    if matchstr:
                        self.update_excel(matchstr,count/total_count*100,'G',counter)
                    else:
                        self.update_excel("Other",count/total_count*100,'H',counter)

                        if mismatch_dic:
                            self.update_excel(mismatch_dic,counter)
                else:
                    # print('not found')
                    # print(f'Sequence:{indexs+1} {sequence} Percentatge:{count/total_count*100:.1f}')
                    self.update_excel("Other",count/total_count*100,'H',counter)                
            else:
                # print('not found')
                # print(f'Sequence:{indexs+1} {sequence} Percentatge:{count/total_count*100:.1f}')
                self.update_excel("Other",count/total_count*100,'H',counter)
            counter += 2

    def check_mutation(self,seq, m_df):
        sorted_df = self.sort_df(m_df)
        column_items = set(sorted_df['mutation'])
        seq_items = set(seq)
        mismatches = column_items.symmetric_difference(seq_items)
        if mismatches:
            # Check if there are extra items in the column or seq
            extra_in_column = column_items - seq_items
            extra_in_seq = seq_items - column_items

            if extra_in_column and extra_in_seq:
                return [f"Missing: {extra_in_column}",f"Extra: {extra_in_seq}"]
            elif extra_in_column:
                return[f"Missing: {extra_in_column}"]
            elif extra_in_seq:
                return[f"Extra: {extra_in_seq}"]
            else:
                return [f"Missmatched: {mismatches}"]

        else:
            return []

    def find_col(self,row):
        for column in range(8, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=column)
                if cell.value is not None:
                    last_column = openpyxl.utils.get_column_letter(column)
        
        next_column = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(last_column) + 1)
        return next_column

    def update_excel(self,*args, **kwargs):
        global sheet
        if len(args) == 1:
            return args[0]
        elif len(args) == 4:
            c1 = sheet[f'{args[2]}{args[3]}']
            c1.value = args[0]
            c2 = sheet[f'{args[2]}{args[3] + 1}']  
            c2.value = args[1]
        elif len(args) == 2:
            col = self.find_col(args[1])
            sorted_dict = dict(sorted(args[0].items(), key=lambda item: len(item[1][0])))
            for key, value in sorted_dict.items():
                c1 = sheet[f'{col}{args[1]}'] 
                c1.value = f'{key}'
                c2 = sheet[f'{col}{args[1]+1}']  
                c2.value = f'{value}'
                col = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(col) + 1)

    def create_excel(self):
        global workbook, sheet

        data = pd.DataFrame()
        data['Label'] = None
        data['Code'] = None
        data['Date'] = None
        data['Algorithm'] = None
        data['Number of Reads'] = None
        data['Sequence'] = None
        data['Match'] = None
        data['Unkown'] = None
        try:
            data.to_excel(result_file.get(),index=False)
            
        except:
            self.logprint("- Error writing output file!")
            messagebox.showerror('Error', 'Error writing output file!') 
            return 1
        self.logprint("Creating Excel file...")
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(result_file.get())

            # Select the active (current) sheet
            sheet = workbook.active        
        except:
            self.logprint("- Error load output file!")
            return 1

    def addto_excel(self,sit,df,file):
        global sheet,workbook
        pattern = r'^.*?15(.*?)(\d+)'
        match = re.search(pattern,sit)
        start_col = 'A'
        if match:
            code = match.group(1)
            date = match.group(2)
            if file.lower().endswith("chim_rm.tsv"):
                algorithm = 'Chimeras_Removed'
            elif file.lower().endswith("covar_deconv.tsv"):
                algorithm = 'Covar_Deconv'
            total_count = df['Count'].sum()
            num_row = len(df['Sequences'])*2
            data=[sit,code,date,algorithm,total_count]
            self.merge_cell(num_row)
            # add data to the sheet
            for item in data:
                cell = sheet[f'{start_col}{sheet.max_row}']  # Get the cell in the specified column and row
                top_left_cell = None
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = sheet.cell(row=merged_range.bounds[1], column=merged_range.bounds[0])
                        break
                top_left_cell.value = item
                # add wrap text
                start_col = chr(ord(start_col) + 1)

    def merge_cell(self,n):
        global workbook, sheet

        # Calculate the starting row for merging (1 row below the current max row)
        start_row = sheet.max_row + 1

        # Define the merge range based on the starting row and the number of rows
        end_row = start_row + n - 1
        for col in range(1, 6):
            column_letter = openpyxl.utils.get_column_letter(col)
            merge_range = f'{column_letter}{start_row}:{column_letter}{end_row}'
            sheet.merge_cells(merge_range)

    def sort_df(self,m_df):
        pattern = r'^S:[A-Z]?(\d+)[A-Z]?$'
        matches = []

        # Loop through each row in the DataFrame
        for index, row in m_df.iterrows():
            mutation = row['mutation']
            
            # find the pattern in the mutation string
            match = re.search(pattern, mutation)
            
            # Check if a match was found and if the numeric part is between 300 and 505
            if match:
                numeric_part = int(match.group(1))
                if 300 <= numeric_part <= 505:
                    matches.append(row)
        matched_df = pd.DataFrame(matches)
        return matched_df
            
    def generateURL(self,lst,url,format):
        api = f'{self.base_url}{url}'
        if(isinstance(lst,list)):
            item_str = ','.join(lst)
        else:
            item_str = lst
        api = f'{api}{item_str}{format}'
        return api

    def parse(self,str):
        matches = re.findall(r'\((.*?)\)', str)
        matches_with_prefix = [f'S:{match}' for match in matches]
        return matches_with_prefix

    def API_call(self,url):
        response = requests.get(url)
        return response

    def result(self):
        global workbook,counter,mutation_url
        country_link = ''
        division_link = ''
        region_link = ''
        if country.get():
            country_link = f'&country={country.get()}'
        if division.get():
            division_link  = f'&division={division.get()}'
        if region.get():
            region_link = f',region'

        mutation_url = f'aggregated?&dateFrom={self.start_date}&dateTo={self.end_date}&fields=pangoLineage{region_link}{country_link}{division_link}&aaMutations='
        inpath = Path(raw_input.get())
        counter = 2
        self.logprint("Adding data to file...")
        # self.logprint("Loading...")
        try:
            ### GET LIST OF VALID INPUT FILES ###
            tsv_files = [
                os.path.join(root, file)
                for root, _, files in os.walk(inpath)
                for file in files
                if (
                    file.lower().endswith("chim_rm.tsv") or
                    file.lower().endswith("covar_deconv.tsv")

                ) and "Collected" not in file
                # if file.lower().endswith(".tsv")

            ]
            # for tsv_file in tsv_files:
                # print("Matching TSV File:", tsv_file)
        except:
            self.logprint("- Error opening directory!")
            messagebox.showerror('Error', 'Error opening directory!')
            return 1

        if not tsv_files:
            self.logprint("- No files were found!")
            messagebox.showwarning('Warning', 'No input files found!')
            return 0
        
        tsv_files.sort()
        self.create_excel()
        for file in tsv_files:      
            try:
                df = pd.read_csv(inpath / file,sep='\t')
                sit = df.columns[0]
                df = pd.read_csv(inpath / file,sep='\t',header=1)
            except:
                self.logprint("- Error opening %s!" % file)
                messagebox.showerror('Error', 'Error opening %s!' % file)
                return 1
            
            for index, row in df.iterrows():
                sequence = row['Sequences']
                pattern = r'\S*fs\S*'
                matches = re.findall(pattern, sequence)
                if matches:
                    df = df.drop(index, axis=0)

            # add output to sheeet
            if not df.empty:
                self.addto_excel(sit,df,file)
                self.handle_df(df)
                print("Loading...")
        workbook.save(result_file.get())
        self.logprint("Done writting to file...")

    def save(self):
        self.end_date = cal2.get_date()
        self.start_date = cal.get_date()
        newWindow.destroy()

    def cancel(self):
        self.set_default()
        newWindow.destroy()

    def set_default(self):
        region.set('')
        country.set('')
        division.set('')


    def setting(self):
        global newWindow,cal,cal2
        newWindow = Toplevel(root)
        newWindow.title("Setting Window")
        row1 = ttk.Frame(newWindow)

        cal=DateEntry(row1,width=12, background='white',
                        foreground='black', borderwidth=2)
        cal.set_date(self.start_date)
        cal.config(state="readonly")

        cal2=DateEntry(row1,width=12, background='white',
                        foreground='black', borderwidth=2)
        cal2.config(state="readonly")
        Label(row1, text="Select start date:").pack(side=LEFT, padx=5)
        cal.pack(side=LEFT, padx=5)
        Label(row1, text="Select end date:").pack(side=LEFT, padx=5)
        cal2.pack(side=LEFT, padx=5)
        row1.pack(side=TOP, padx=5, pady=5)

        row2 = ttk.Frame(newWindow)
        Label(row2,text="Choose a region").pack(side=LEFT, padx=5)
        drop1 = OptionMenu(row2,region, *self.world_region)
        drop1.pack(side=LEFT,pady=5)
        Label(row2,text="Choose a country").pack(side=LEFT, padx=5)
        entry1 = ttk.Entry(row2,width=20,textvariable=country)
        entry1.pack(side=RIGHT, expand=YES, fill=X)
        row2.pack(side=TOP, padx=5, pady=5)

        row3 = ttk.Frame(newWindow)
        Label(row3,text="Choose a division").pack(side=LEFT, padx=5)
        drop2 = OptionMenu(row3,division, *self.us_states)
        drop2.pack(side=LEFT,pady=5)
        Label(row3,text="Enter a threshold").pack(side=LEFT, padx=5)
        entry2 = ttk.Entry(row3,width=20,textvariable=threshold)
        entry2.pack(side=RIGHT, expand=YES, fill=X)
        row3.pack(side=TOP, padx=5, pady=5)

        row4 = ttk.Frame(newWindow)
        ttk.Button(row4, text="Save", command=self.save).pack(side=LEFT,padx=15)
        ttk.Button(row4, text="Cancel", command=self.cancel).pack(side=LEFT,padx=15)
        row4.pack(side=TOP, padx=5, pady=5)

        self.set_default()

if __name__ == '__main__':
    root = Tk()
    app = VariantSpotterApp(root)
    root.mainloop()